import csv
import math
import sys
import os
import zipfile

import openpyxl
import pandas as pd
import psutil
import time as t
from PyQt5 import QtWidgets, uic
from PyQt5.QtCore import pyqtSlot, Qt, QEvent, QSize, QObject
from PyQt5.QtWidgets import QTableWidgetItem, QSizePolicy, QVBoxLayout, QLayout, QApplication, QWidget, QHBoxLayout, \
    QLabel, QComboBox, QPushButton, QCheckBox
from openpyxl.styles import Alignment, PatternFill


class ScalingEventFilter(QObject):
    def __init__(self):
        super().__init__()
        self.pre=1
    def eventFilter(self, obj, event):
        #only works on windows, requires extra code to get dpi if want to work on mac
        if event.type() == QEvent.Resize and obj.isWindow():

            if sys.platform == 'win32':
                scaling_factor = obj.logicalDpiX() / 96.0
            elif sys.platform == 'darwin':
                # On macOS, you need to get the DPI from the screen
                screen = obj.windowHandle().screen()
                scaling_factor = screen.logicalDotsPerInch() / 96.0

            contract = scaling_factor
            # if scaling_factor > 2.25:
            #
            #     scaling_factor=2.25


            if self.pre==1 or self.pre!=scaling_factor:
                print(scaling_factor)
                print('scaling change event happened')

                #obj.resize(1500/self.pre*scaling_factor,800/self.pre*scaling_factor)
                if scaling_factor is not  1.5:
                    desired_font_size = 8 /scaling_factor*1.5# Set a fixed font size of 8

                else:
                    desired_font_size = 8
                for widget in obj.findChildren(QWidget):
                    #if isinstance(widget, QLabel):
                        # Scale font size based on scaling factor
                    font = widget.font()
                    font.setPointSizeF( desired_font_size )
                    widget.setFont(font)


                self.pre=scaling_factor



            #return super().eventFilter(obj, event)



            # Adjust the frame's size based on the custom scaling factor




        return super().eventFilter(obj, event)




class Ui(QtWidgets.QMainWindow):

    def __init__(self):
        self.previous_value = 0
        self.file=0
        #this statement is important
        self.file2=1
        self.count=1
        self.check2=True

        app = QtWidgets.QApplication(sys.argv)
        steel_num=['3#','4#','5#','6#','7#','8#','9#','10#','11#']
        steel_type=['280','280w','420','420w','490w']
        length=['12','14','15','16','18']

        self.window = uic.loadUi("outline3.ui")
        os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1.0"




        #os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"

        self.window.pushButton.clicked.connect(lambda:self.button1())

        self.window.pushButton_2.clicked.connect(lambda:self.button2())
        self.window.pushButton_3.clicked.connect(lambda:self.button3())
        self.window.pushButton_4.clicked.connect(lambda: self.button4())
        self.window.comboBox.addItem("A")

        self.window.comboBox_2.addItems(steel_num)
        self.window.comboBox_3.addItems(steel_type)
        self.window.comboBox_4.addItems(length)

        self.buildTextEditList()
        self.update_TextEdit()
        self.existed_input()
        self.currentFile="File1"

        self.window.comboBox.currentTextChanged.connect(lambda:self.update_TextEdit(self.window.comboBox.currentText()))
        self.window.spinBox.setRange(1, 300)
        self.buildspinnerList()
        self.window.tableWidget.setRowCount(7)
        self.window.tableWidget.setColumnCount(2)
        self.initial=False
        self.existed_file()
        self.window.pushButton_5.clicked.connect(lambda: self.append_df_to_excel(self.currentFile))
        self.window.pushButton_6.clicked.connect(lambda: self.MachineList())
        self.window.pushButton_8.clicked.connect(lambda: self.steal_Save())
        self.update_table(self.currentFile)
        self.window.spinBox.valueChanged.connect(lambda:self.main_data(self.currentFile))

        self.window.pushButton_7.clicked.connect(lambda: self.on_combobox_changed2(self.currentFile))
        self.read()


        #self.main_data()
        # Set the item at row 0 and column 0
        event_filter = ScalingEventFilter()

        self.window.installEventFilter(event_filter)


        self.window.show()
        app.exec()
    def MachineList(self):
        try:
            with open("save_data"+str(self.window.comboBox.currentText())+".txt", 'r') as f:
                csvreader = csv.reader(f)
            index = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
            print(index[self.window.comboBox.count()])
            print(self.window.comboBox.currentText())
            if index[self.window.comboBox.count()-1]==self.window.comboBox.currentText():

                self.window.comboBox.addItem(index[self.window.comboBox.count()])
        except FileNotFoundError:
            print('failed')


    def on_combobox_changed2(self, value):


            print("entered"+value)
            self.window.tableWidget.clearContents()
            #self.main_data(value)
            #self.update_table(value)
            print(eval(value[-1:])+1)
            self.file2+=1
            self.currentFile = ("File" + str(self.file2 - 1))
            self.check2=False
            self.window.spinBox.setValue(1)
            self.check2=True


            df_empty = pd.DataFrame()

            df_empty.to_csv(self.currentFile  +"System.csv", index=False,header=False)
    def existed_input(self):

        try:
            index = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']
            self.file += 1
            print("save_data" + str(index[self.file]))

            with open("save_data" + str(index[self.file])+".txt", 'r') as f:
                csvreader = csv.reader(f)

            self.window.comboBox.addItem(index[self.file])

            self.existed_input()


        except FileNotFoundError:

            print('bye')

    def existed_file(self):

        try:

            self.file2 += 1
            print("File" + str(self.file2)+"System.csv")

            with open("File" + str(self.file2)+"System.csv", 'r') as f:
                csvreader = csv.reader(f)



            self.existed_file()


        except FileNotFoundError:

            self.currentFile=("File" + str(self.file2-1))

    def calcWeight(self,length,amount):
        list1 = length
        list2 = amount
        dict1 = {'3#': 0.56, '4#': 0.994, '5#': 1.56, '6#': 2.25, '7#': 3.04, '8#': 3.98, '9#': 5.08, '10#': 6.39,
                 '11#': 7.9}
        # placeholder
        i = 0
        result = 0

        while i in range(len(list1)):

            if self.is_num(list1[i]) and self.is_num(list2[i]):
                result += round(int(list1[i]) / 100 * int(list2[i]) * dict1[self.window.comboBox_2.currentText()])
            i += 1
        print(result)
        return(result)




    def append_df_to_excel(self,fileName,initial=False):

        try:


          
            index = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']
            i = 0
            if not self.initial :


                    #self.count=1
                if not self.inCsv(self.window.spinBox.value(),fileName):
                    self.specialAppend(fileName)
                os.remove(fileName+".xlsx")

                print('here')



            print("bruh")
            writer = pd.ExcelWriter(fileName + ".xlsx", engine='openpyxl',  mode='a', if_sheet_exists='overlay')
            weight=[]
            templen=0
            num_rol=0

            with open(fileName+"System.csv", 'r', newline='') as file:
                print("enter")
                reader = csv.reader(file)

                rows = list(reader)

            count=0

            temp=[]
            for x in rows:

                if not x[0].__contains__('Steel'):

                    if count%2==0:
                        df2 = pd.DataFrame({"Length":x})
                        templen=x
                    elif count%2==1:
                        df2=pd.DataFrame({"Amount":x})
                        print(templen)
                        print(x)
                        weight+=[self.calcWeight(templen,x),""]
                    num_rol=((math.floor((count)/26))*10+1)
                    num_col=count-((math.floor((count)/26))*25+math.floor((count)/26))
                    print('dsfsdf')
                    df2.to_excel(writer, sheet_name="Sheet1", index=False, startrow=num_rol, startcol=num_col)
                    count+=1
                else:
                    temp += [x[0]]
                    temp+=[""]


            k = 0
            print(weight)


            while k in range(num_rol):

                if len(temp)>=((math.ceil((k+1)/10))*26)-1:

                    df = pd.DataFrame(temp[math.floor(k/10)*26:((math.ceil((k+1)/10))*26)-1])

                else:

                    df = pd.DataFrame(temp[math.floor(k/10)*26:len(temp)])
                if len(weight)>=((math.ceil((k+1)/10))*26)-1:
                    df2 = pd.DataFrame(weight[math.floor(k / 10) * 26:((math.ceil((k + 1) / 10)) * 26) - 1])
                else:
                    df2 = pd.DataFrame(weight[math.floor(k / 10) * 26:len(temp)])
                df=df.T

                df2=df2.T

                df.to_excel(writer, sheet_name="Sheet1", index=False, header=False,startrow=k, startcol=0)

                df2.to_excel(writer, sheet_name="Sheet1", index=False, header=False,startrow=k+9, startcol=0)


                while i in range(len(index) - 1):
                    worksheet = writer.sheets['Sheet1']
                    worksheet.merge_cells(index[i] + str(k + 10) + ':' + index[i + 1] + str(k + 10))
                    worksheet.merge_cells(index[i] + str(k+1) + ':' + index[i + 1] + str(k+1))

                    i += 2
                i=0

                k += 10


            # for column_idx in range(len(df.columns)):
            #     for cell in worksheet[openpyxl.utils.cell.get_column_letter(column_idx + 1)]:
            #         cell.alignment = Alignment(horizontal="center")
            # center all cells

            for col in worksheet.columns:
                for cell in col:
                    # Create a new Alignment object with the desired alignment settings
                    alignment_obj = Alignment(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj


            writer.close()

            self.initial=False
        except(FileNotFoundError, zipfile.BadZipFile):
            self.initial=True
            df_empty = pd.DataFrame()

            df_empty.to_excel(fileName+".xlsx", index=False)

            self.append_df_to_excel(fileName,self.initial)



    def change(self):
        current_value = self.window.spinBox.value()

        if current_value > self.previous_value:
            self.previous_value = current_value
            #print('greater')
            return True
        elif current_value < self.previous_value:

            self.previous_value = current_value
            return False
        else:
            #print('equal')
            return 'equal'

    def specialAppend(self,fileName):
        print("check")
        row = 0
        row2 = 0
        steel_length=[]
        steel_amount=[]

        while row in range(self.window.tableWidget.rowCount()) and self.window.tableWidget.item(row, 0) != None:
            steel_length += [self.window.tableWidget.item(row, 0).text()]
            row += 1
        while row2 in range(self.window.tableWidget.rowCount()) and self.window.tableWidget.item(row2, 1) != None:
            steel_amount += [self.window.tableWidget.item(row2, 1).text()]
            row2 += 1
        with open(fileName + "System.csv", 'a', newline='') as file:


            i2 = str(self.window.spinBox.value())
            writer = csv.writer(file)

            writer.writerow(["Steel" + i2])
            writer.writerow(steel_length)

            writer.writerow(steel_amount)


    def main_data(self,fileName):
        try:



            self.steel_length=[]
            self.steel_amount=[]

            row=0
            row2=0


            while row in range(self.window.tableWidget.rowCount()) and self.window.tableWidget.item(row, 0)!=None:
                temp=[self.window.tableWidget.item(row, 0)]
                for x in temp:
                    if self.is_num(x.text()):
                        self.steel_length+=[x.text()]
                row+=1
            while row2 in range(self.window.tableWidget.rowCount()) and self.window.tableWidget.item(row2, 1)!=None:


                temp = [self.window.tableWidget.item(row2, 1)]
                for x in temp:
                    if self.is_num(x.text()):
                        self.steel_amount+=[x.text()]
                row2 += 1
            if len(self.steel_amount)==0 or len(self.steel_length)==0:
                pass;
            else:

                print(fileName+"System.csv")
                if self.check2 and not self.inCsv(self.window.spinBox.value(),fileName):


                    with open(fileName+"System.csv", 'a', newline='') as file:
                        if self.change():
                            i=str(self.window.spinBox.value()-1)
                        else:
                            i=str(self.window.spinBox.value()+1)
                        writer = csv.writer(file)

                        writer.writerow(["Steel"+i])
                        writer.writerow(self.steel_length)
                        writer.writerow(self.steel_amount)


                else:


                    if len(self.steel_length)>0 or len(self.steel_amount)>0:
                        result=self.change()
                        if result==True:

                            target2 = ( 3 * (int(self.window.spinBox.value()) - 2))
                        elif result=='equal':
                            target2=( 3 * (int(self.window.spinBox.value())-1 ))
                        else:
                            target2=( 3 * (int(self.window.spinBox.value()) ))

                        self.replace_row(fileName+"System.csv",target2,self.steel_length,self.steel_amount,True)


            self.update_table(fileName)


            #self.change()




        except FileNotFoundError:
            with open(fileName+'System.csv', 'a', newline='') as file:
                writer = csv.writer(file)

                self.main_data(fileName)
    def update_table(self,fileName):
        try:
            currentLine = 1
            with open(fileName+"System.csv", 'r') as f:
                csvreader = csv.reader(f)

                target = (currentLine + 3 * (int(self.window.spinBox.value()) - 1))


                interestingrow = [row for idx, row in enumerate(csvreader) if idx == target]

                f.seek(0)
                interestingrow2 = [row2 for idx1, row2 in enumerate(csvreader) if idx1 == target + 1]

                setrow = 0
                if len(interestingrow) == 0 or len(interestingrow2) == 0:

                    self.window.tableWidget.clearContents()
                else:

                    while setrow in range(self.window.tableWidget.rowCount()):
                        if len(interestingrow[0]) > setrow:
                            item = QTableWidgetItem(interestingrow[0][setrow])
                            self.window.tableWidget.setItem(setrow, 0, item)
                        else:
                            ("bruh")
                            self.window.tableWidget.setItem(setrow, 0, QTableWidgetItem(""))
                        setrow += 1
                    setrow2 = 0
                    while setrow2 in range(self.window.tableWidget.rowCount()):

                        if len(interestingrow2[0]) > setrow2:
                            item = QTableWidgetItem(interestingrow2[0][setrow2])
                            self.window.tableWidget.setItem(setrow2, 1, item)
                        else:
                            # print("bruh")
                            self.window.tableWidget.setItem(setrow2, 1, QTableWidgetItem(""))
                        setrow2 += 1
        except FileNotFoundError:
            print("ok")

    def replace_row(self,csv_file, index, new_row,new_row2,head):
        with open(csv_file, 'r', newline='') as file:
            reader = csv.reader(file)

            rows = list(reader)

        # Replace the desired row with the new row
        if index>=len(rows)-1:

            if head:
                rows.append(["Steel"+str(self.window.spinBox.value()+1)])

                rows += [new_row]
                rows += [new_row2]
            else:
                rows += [new_row]
                rows += [new_row2]
        else:


            rows[index+1]=new_row
            rows[index+2]=new_row2



        # Write the updated data back to the CSV file
        with open(csv_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

    def inCsv(self,key1,fileName):

        with open(fileName+"System.csv", 'r', newline='') as file:

            csvreader = csv.reader(file)

            if os.path.getsize(fileName+"System.csv") == 0:

                return False
            for row in csvreader:

                if row[0] == "Steel" + str(key1) or row[0] == "Steel" + str(key1-1):

                    return True
            return False


    def build_data(self):
        self.steel_data=[self.window.comboBox_2.currentText(),self.window.comboBox_3.currentText(),self.window.comboBox_4.currentText(),self.window.checkBox.isChecked(),self.window.textEdit_20.toPlainText()]
        self.prelength=[self.window.textEdit_21.toPlainText(),self.window.textEdit_22.toPlainText(),self.window.textEdit_23.toPlainText(),self.window.textEdit_24.toPlainText(),self.window.textEdit_25.toPlainText()]
    def button2(self):
        self.build_data()
        self.pausing(False)


    def button3(self):
        self.pausing(True)
        print('paused')
    def button4(self):
        os.kill()
    def pausing(self,program):
        print("suppose to be pausing")

    def buildTextEditList(self):
        self.TextEdits=[]
        self.TextEdits += [self.window.A1]
        self.TextEdits += [self.window.A2]
        self.TextEdits += [self.window.A3]
        self.TextEdits += [self.window.A4]
        self.TextEdits += [self.window.A5]
        self.TextEdits += [self.window.A6]
        self.TextEdits += [self.window.A7]
        self.TextEdits += [self.window.A8]
        self.TextEdits += [self.window.A9]
        self.TextEdits += [self.window.A10]
        self.TextEdits += [self.window.A11]
        self.TextEdits += [self.window.A12]
        self.TextEdits += [self.window.A13]
        self.TextEdits += [self.window.A14]
        self.TextEdits += [self.window.A15]
        self.TextEdits += [self.window.A16]
        self.TextEdits += [self.window.A17]
        self.TextEdits += [self.window.A18]
        self.TextEdits += [self.window.A19]
        self.TextEdits += [self.window.A20]
        self.TextEdits += [self.window.A21]
        self.TextEdits += [self.window.A22]
        self.TextEdits += [self.window.A23]
        self.TextEdits += [self.window.A24]
        self.TextEdits += [self.window.textEdit_11]
        self.TextEdits += [self.window.textEdit_12]
        self.TextEdits += [self.window.textEdit_13]
        self.TextEdits += [self.window.textEdit_14]
        self.TextEdits += [self.window.textEdit_15]
        self.TextEdits += [self.window.textEdit_16]
        self.TextEdits += [self.window.textEdit_17]
        self.TextEdits += [self.window.textEdit_18]
        self.TextEdits += [self.window.textEdit_19]
    def buildspinnerList(self):
        self.spinner=[]
        self.window.B1.setRange(-30, 30)
        self.window.B2.setRange(-30, 30)
        self.window.B3.setRange(-30, 30)
        self.window.B4.setRange(-30, 30)
        self.window.B5.setRange(-30, 30)
        self.window.B6.setRange(-30, 30)
        self.window.B7.setRange(-30, 30)
        self.window.B8.setRange(-30, 30)
        self.window.B9.setRange(-30, 30)
        self.window.B10.setRange(-30, 30)
        self.window.B11.setRange(-30, 30)
        self.window.B12.setRange(-30, 30)
        self.window.B13.setRange(-30, 30)
        self.window.B14.setRange(-30, 30)


        self.spinner += [self.window.B1]
        self.spinner += [self.window.B2]
        self.spinner += [self.window.B3]
        self.spinner += [self.window.B4]
        self.spinner += [self.window.B5]
        self.spinner += [self.window.B6]
        self.spinner += [self.window.B7]
        self.spinner += [self.window.B8]
        self.spinner += [self.window.B9]
        self.spinner += [self.window.B10]
        self.spinner += [self.window.B11]
        self.spinner += [self.window.B12]
        self.spinner += [self.window.B13]
        self.spinner += [self.window.B14]




    def steal_Save(self):

        self.limits=[]
        for x in self.spinner:
            self.limits+=[str(x.value())]
        file = "limits.txt"
        append = open(file, "w")
        print(self.limits)

        append.writelines("\n".join(self.limits))
    def read(self):
        read = open("limits.txt", "r")
        for spin in self.spinner:
            spin.setValue(int(read.readline()))




    def on_combobox_changed(self, value):
        print("fucked u mate")
        self.update_TextEdit(value)

        # do your code
    def is_num(self,value):
        try:
            int(value)
            return True
        except :
            return False
    def update_TextEdit(self, value1='A'):
        try:


            index=0
            fileName="save_data"+value1+".txt"
            with open(fileName,'r') as read:
                line=read.readline().strip()

                while line!='end':


                    if(self.is_num(line)or line==''):

                        self.TextEdits[index].setPlainText(line)
                        index += 1
                    else:
                        print("bruh ur stupid")
                    line=read.readline().strip()




        except FileNotFoundError:

           for x in self.TextEdits:
               x.setPlainText("")

    def button1(self):

        self.data()
        if self.length_limit[0]!='end':
            file="save_data"+self.window.comboBox.currentText()+".txt"
            new=open(file,"w")

            new.writelines("\n".join(self.constraint))
            new.write("\n")

            new.writelines("\n".join(self.length_limit))
        else:
            print('here')
            self.update_TextEdit(self.window.comboBox.currentText())













    def data(self):
        self.constraint=[]
        self.length_limit=[]
        check=False
        print(len(self.TextEdits))
        for i in range(24):
            print((self.TextEdits[i].toPlainText()))
            if not self.is_num(self.TextEdits[i].toPlainText()) and self.TextEdits[i].toPlainText().strip()!='':
                check = True
                break

            else:

                self.constraint+=[self.TextEdits[i].toPlainText()]
            i+=1

        for k in range(24,33):
            #for test only
            if check or (not self.is_num(self.TextEdits[k].toPlainText()) and self.TextEdits[k].toPlainText().strip()!=''):

                self.constraint.clear()
                self.length_limit.clear()
                break
            else:
                print(self.TextEdits[25].toPlainText())
                print('up')

                self.length_limit += [self.TextEdits[k].toPlainText()]
                print('here')
            k += 1
        print(self.constraint)
        print(self.length_limit)
        self.length_limit+=['end']





Ui()